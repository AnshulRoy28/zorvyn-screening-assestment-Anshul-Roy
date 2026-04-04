import base64
import csv
from io import StringIO
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from app.middleware.auth import require_role
from app.services.gemini_service import (
    extract_from_receipt, extract_from_bank_statement, classify_excel_rows
)
from app.services.csv_parser import parse_transactions_from_statement

upload_bp = Blueprint('upload', __name__, url_prefix='/upload')

@upload_bp.route('/receipt', methods=['POST'])
@require_role('admin')
def upload_receipt():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    try:
        image_data = base64.b64encode(file.read()).decode('utf-8')
        transactions, error = extract_from_receipt(image_data)
        
        if error:
            return jsonify({'success': False, 'error': error}), 500
        
        for t in transactions:
            t['source'] = 'receipt'
        
        return jsonify({
            'success': True,
            'data': {
                'transactions': transactions,
                'message': 'Preview extracted. Review and confirm to save.'
            }
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Processing error: {str(e)}'}), 500

@upload_bp.route('/bank-statement', methods=['POST'])
@require_role('admin')
def upload_bank_statement():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    try:
        image_data = base64.b64encode(file.read()).decode('utf-8')
        transactions, error = extract_from_bank_statement(image_data)
        
        if error:
            return jsonify({'success': False, 'error': error}), 500
        
        for t in transactions:
            t['source'] = 'bank_statement'
        
        return jsonify({
            'success': True,
            'data': {
                'transactions': transactions,
                'message': 'Preview extracted. Review and confirm to save.'
            }
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Processing error: {str(e)}'}), 500

@upload_bp.route('/excel', methods=['POST'])
@require_role('admin')
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    filename_lower = file.filename.lower()
    is_csv = filename_lower.endswith('.csv')
    is_excel = filename_lower.endswith(('.xlsx', '.xls'))
    
    if not (is_csv or is_excel):
        return jsonify({'success': False, 'error': 'File must be CSV or Excel format'}), 400
    
    try:
        rows = []
        
        if is_csv:
            # Parse CSV file
            content = file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
            
            # Try ICICI bank statement parser first
            try:
                rows = parse_transactions_from_statement(content)
            except Exception:
                # Fallback to generic CSV parser
                csv_reader = csv.reader(StringIO(content))
                all_rows = list(csv_reader)
                
                if not all_rows:
                    return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
                
                # Find header row
                headers = []
                data_start_row = 0
                
                for idx, row in enumerate(all_rows[:10]):
                    if row and any(row):
                        row_str = ' '.join(str(cell).lower() for cell in row if cell)
                        if any(keyword in row_str for keyword in ['date', 'description', 'amount', 'debit', 'credit', 'transaction', 'particulars']):
                            headers = [str(cell).lower() if cell else '' for cell in row]
                            data_start_row = idx + 1
                            break
                
                if not headers and len(all_rows) > 1:
                    data_start_row = 1
                
                rows = parse_rows(all_rows[data_start_row:], headers)
        
        else:
            # Parse Excel file
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active
            
            headers = []
            data_start_row = 2
            
            # Find headers in first few rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(row):
                    row_str = ' '.join(str(cell).lower() for cell in row if cell)
                    if any(keyword in row_str for keyword in ['date', 'description', 'amount', 'debit', 'credit', 'transaction', 'particulars']):
                        headers = [str(cell).lower() if cell else '' for cell in row]
                        data_start_row = row_idx + 1
                        break
            
            # Get all data rows
            all_rows = []
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                if row and any(row):
                    all_rows.append(row)
            
            rows = parse_rows(all_rows, headers)
        
        if not rows:
            return jsonify({'success': False, 'error': 'No valid data found. Please ensure file has Date, Description, and Amount columns.'}), 400
        
        # Classify with Gemini
        descriptions = [r['description'] for r in rows]
        classifications, error = classify_excel_rows(descriptions)
        
        if error:
            return jsonify({'success': False, 'error': error}), 500
        
        transactions = []
        for i, row in enumerate(rows):
            if i < len(classifications):
                transactions.append({
                    'item_name': classifications[i].get('item_name', row['description']),
                    'amount': row['amount'],
                    'category': classifications[i].get('category', 'Other'),
                    'type': classifications[i].get('type', row['type']),
                    'date': row['date'],
                    'source': 'excel'
                })
        
        return jsonify({
            'success': True,
            'data': {
                'transactions': transactions,
                'message': 'Preview extracted. Review and confirm to save.'
            }
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Processing error: {str(e)}'}), 500


def parse_rows(data_rows, headers):
    """Parse rows from CSV or Excel data"""
    rows = []
    
    # Find column indices
    date_col = None
    desc_col = None
    amount_col = None
    debit_col = None
    credit_col = None
    particulars_col = None
    deposits_col = None
    withdrawals_col = None
    
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if 'date' in header_lower:
            date_col = idx
        elif 'particulars' in header_lower or 'description' in header_lower or 'narration' in header_lower or 'details' in header_lower:
            desc_col = idx
            particulars_col = idx
        elif 'deposits' in header_lower or 'credit' in header_lower or 'deposit' in header_lower:
            deposits_col = idx
            credit_col = idx
        elif 'withdrawals' in header_lower or 'debit' in header_lower or 'withdrawal' in header_lower:
            withdrawals_col = idx
            debit_col = idx
        elif 'amount' in header_lower and 'debit' not in header_lower and 'credit' not in header_lower:
            amount_col = idx
    
    # Parse data rows
    for row in data_rows:
        if not row or not any(row):
            continue
        
        # Skip rows that look like headers or summaries
        if len(row) > 0:
            first_cell = str(row[0]).lower()
            if any(skip_word in first_cell for skip_word in ['date', 'total', 'balance', 'account', 'statement', 'b/f', 'c/f']):
                continue
        
        try:
            # Extract date
            date_val = None
            if date_col is not None and len(row) > date_col:
                date_val = row[date_col]
            elif len(row) > 0:
                date_val = row[0]
            
            # Skip if no date
            if not date_val or str(date_val).strip() == '':
                continue
            
            # Extract description (PARTICULARS column)
            description = ''
            if particulars_col is not None and len(row) > particulars_col:
                description = str(row[particulars_col]) if row[particulars_col] else ''
            elif desc_col is not None and len(row) > desc_col:
                description = str(row[desc_col]) if row[desc_col] else ''
            elif len(row) > 2:  # Try third column if no header match
                description = str(row[2]) if row[2] else ''
            elif len(row) > 1:
                description = str(row[1]) if row[1] else ''
            
            # Skip empty descriptions or common non-transaction entries
            if not description or description.lower() in ['none', 'null', '', 'b/f', 'c/f']:
                continue
            
            # Extract amount from DEPOSITS or WITHDRAWALS columns
            amount = None
            trans_type = 'expense'
            
            # Try deposits column (income)
            if deposits_col is not None and len(row) > deposits_col and row[deposits_col]:
                try:
                    deposits_val = str(row[deposits_col]).replace(',', '').strip()
                    if deposits_val and deposits_val != '0' and deposits_val != '0.00':
                        amount = float(deposits_val)
                        trans_type = 'income'
                except (ValueError, TypeError):
                    pass
            
            # Try withdrawals column (expense)
            if withdrawals_col is not None and len(row) > withdrawals_col and row[withdrawals_col]:
                try:
                    withdrawals_val = str(row[withdrawals_col]).replace(',', '').strip()
                    if withdrawals_val and withdrawals_val != '0' and withdrawals_val != '0.00':
                        amount = float(withdrawals_val)
                        trans_type = 'expense'
                except (ValueError, TypeError):
                    pass
            
            # Fallback to debit/credit columns
            if amount is None and debit_col is not None and len(row) > debit_col and row[debit_col]:
                try:
                    amount = float(str(row[debit_col]).replace(',', ''))
                    trans_type = 'expense'
                except (ValueError, TypeError):
                    pass
            
            if amount is None and credit_col is not None and len(row) > credit_col and row[credit_col]:
                try:
                    amount = float(str(row[credit_col]).replace(',', ''))
                    trans_type = 'income'
                except (ValueError, TypeError):
                    pass
            
            # Fallback to amount column
            if amount is None and amount_col is not None and len(row) > amount_col:
                try:
                    amount = abs(float(str(row[amount_col]).replace(',', '')))
                except (ValueError, TypeError):
                    pass
            
            # Skip if no valid amount or amount is 0
            if amount is None or amount == 0:
                continue
            
            rows.append({
                'date': str(date_val) if date_val else None,
                'description': description[:200],  # Limit description length
                'amount': amount,
                'type': trans_type
            })
        except Exception:
            # Skip problematic rows
            continue
    
    return rows